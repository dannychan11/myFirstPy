import  xlrd as xlsr
from datetime import datetime,date,time
import  xlwt as xlsw
import openpyxl as  advRWE
# c1=sheet.cell(sheet.max_row+1,1,'100')
# wb.save(r'D:\11.xlsx')

'''
操作Excel
'''
class operExcel:

    def data2ArrFor2003(data,exl_sheet):
        '''
        将excel封装成一个二位数组,暂时只支持xls.
        :param data excel的文件对象，用来解决Date的问题
        :type workbook
        :param exl_sheet exceld sheet对象 ，必须要formatting_info=True
        :type sheet
        '''
        xlsArr = []
        tempCell = {}
        for merged_cell in exl_sheet.merged_cells:
            min_row, max_row, min_col, max_col = merged_cell
            for firstStr in range(min_row, max_row ):
                for sedStr in range(min_col, max_col):
                    tempCell[str(firstStr) + str(sedStr)] = \
                        exl_sheet.cell(min_row, min_col).value
        for row_index in range(exl_sheet.nrows):
            rowArr = []
            for col_index in range(exl_sheet.ncols):
                '''这里判断当前单元格是否为空，为空是就取上个单元格的值，如果上一个也没值就取上上个，
                主要解决合并单元格的问题，但单元格本身为空时无法解决，下一次解决，xls是可以解决的
                ，已经解决'''
                if exl_sheet.cell(row_index,col_index).value=='':
                    if str(row_index)+str(col_index) in tempCell.keys():
                        rowArr.append(tempCell.get(str(row_index)+str(col_index)))
                    else:
                        rowArr.append('')
                else:
                    if exl_sheet.cell(row_index,
                                col_index).ctype==3:
                        date_value=xlsr.xldate_as_tuple(exl_sheet.cell(row_index,col_index).value,data.datemode)
                        date_tmp=datetime(*date_value).strftime('%Y/%m/%d %H:%M:%S')
                        rowArr.append(date_tmp)
                    else:
                        rowArr.append(exl_sheet.cell(row_index,
                                              col_index).value)
            xlsArr.append(rowArr)
        return xlsArr

    def data2ArrFor2007(exl_sheet):
        '''
        将excel封装成一个二位数组,暂时只支持xlsx.
        :param exl_sheet exceld sheet对象
        :type sheet
        '''
        #将合并单元格的格子都赋值成一个值，先存到一个字典中
        tempCell = {}
        for mergCells in exl_sheet.merged_cells:
            for firstStr in range(mergCells.min_row, mergCells.max_row + 1):
                for sedStr in range(mergCells.min_col, mergCells.max_col + 1):
                    tempCell[str(firstStr) + str(sedStr)] = \
                        exl_sheet.cell(mergCells.min_row, mergCells.min_col).value
        # 计算出第一行的列数，按此列数循环
        counter = 0
        for firRow in exl_sheet.rows:
            for a in firRow:
                counter += 1
            break
        xlsArr = []
        for row_index in range(exl_sheet.max_row):
            rowArr = []
            #只取第一行的列数
            for clo_index in range(counter):
                #下标从一开始所以都要加1
                if exl_sheet.cell(row_index + 1, clo_index + 1).value is None:
                    if str(row_index) + str(row_index + 1) in tempCell.keys():
                        rowArr.append(tempCell[str(row_index) + str(row_index + 1)])
                    else:
                        rowArr.append('')
                else:
                    if isinstance(exl_sheet.cell(row_index + 1, clo_index + 1).value, datetime):
                        rowArr.append(exl_sheet.cell(row_index + 1, clo_index + 1).value.strftime('%Y/%m/%d %H:%M:%S'))
                    else:
                        rowArr.append(exl_sheet.cell(row_index + 1, clo_index + 1).value)
            xlsArr.append(rowArr)
        return xlsArr

    # #在最后一行加一列
    # def writeOneRow(self,exl_sheet,valuesArr):
    #     exl_sheet


