"""日模块将所有操作excel的方法进行提供，并提供main方法调试"""

from datetime import datetime

import xlrd
from xlutils.copy import copy


def readexcelfor2003(workbook_, sheet_):
    """
    将excel封装成一个二位数组,暂时只支持xls.

    :param workbook_: excle的文件对象，用来解决日期问题的问题

    :param sheet_: excel sheet对象 ，必须要formatting_info=True

    :return xlslist: 返回一个二维数组

    """
    xlslist = []  # 存储最终结果集
    tempcell = {}  # 存储合并单元格的位置与值的字典
    for merged_cell in sheet_.merged_cells:  # 将合并单元格的数据都赋值成同一个
        min_row, max_row, min_col, max_col = merged_cell
        for startrow_index in range(min_row, max_row):
            for endcol_index in range(min_col, max_col):
                tempcell[str(startrow_index) + str(endcol_index)] = sheet_.cell(min_row, min_col).value
    for row_index in range(sheet_.nrows):
        rowlist = []  # 存储每一行的数据
        for col_index in range(sheet_.ncols):
            """ 这里判断当前单元格是否为空，为空是就取上个单元格的值，如果上一个也没值就取上上个，
            主要解决合并单元格的问题，但单元格本身为空时无法解决，下一次解决，xls是可以解决的。
            2018-4-29，肖志君修改将2003版本与2007版本分开读写都是               
            """
            if sheet_.cell(row_index, col_index).value == '':
                if str(row_index)+str(col_index) in tempcell.keys():
                    rowlist.append(tempcell.get(str(row_index)+str(col_index)))
                else:
                    rowlist.append('')
            else:
                if sheet_.cell(row_index, col_index).ctype == 3:
                    date_value = xlrd.xldate_as_tuple(sheet_.cell(row_index,
                                                                  col_index).value, workbook_.datemode)
                    date_tmp = datetime(*date_value).strftime('%Y/%m/%d %H:%M:%S')
                    rowlist.append(date_tmp)
                else:
                    rowlist.append(sheet_.cell(row_index, col_index).value)
        xlslist.append(rowlist)
    return xlslist


def readexcelfor2007(sheet_):
    """
    将excel封装成一个二位数组,暂时只支持xlsx.

    :param sheet_: excel的sheet对象

    :return: 二维数组

    """

    tempcell = {}  # 将合并单元格的格子都赋值成一个值，先存到一个字典中
    for mergCells in sheet_.merged_cells:
        for firstStr in range(mergCells.min_row, mergCells.max_row + 1):
            for sedStr in range(mergCells.min_col, mergCells.max_col + 1):
                tempcell[str(firstStr) + str(sedStr)] = \
                    sheet_.cell(mergCells.min_row, mergCells.min_col).value

    clonum = 0  # 计算出第一行的列数，按此列数循环
    for row_ in sheet_.rows:
        clonum = len(row_)
        break
    xlslist = []
    for row_index in range(sheet_.max_row):
        rowlist = []  # 只取第一行的列数
        for clo_index in range(clonum):
            if sheet_.cell(row_index + 1, clo_index + 1).value is None:  # 下标从一开始所以都要加1
                if str(row_index) + str(row_index + 1) in tempcell.keys():
                    rowlist.append(tempcell[str(row_index) + str(row_index + 1)])
                else:
                    rowlist.append('')
            else:
                if isinstance(sheet_.cell(row_index + 1, clo_index + 1).value, datetime):
                    rowlist.append(sheet_.cell(row_index + 1, clo_index + 1).value.strftime('%Y/%m/%d %H:%M:%S'))
                else:
                    rowlist.append(sheet_.cell(row_index + 1, clo_index + 1).value)
        xlslist.append(rowlist)
    return xlslist


def writerowfor2007(workbook_, filepath_, sheet_, valuelist):
    """
    在excel文件的最后插入一行数据.

    :param workbook_: excel object

    :param filepath_: 文件路径

    :param sheet_: sheet object

    :param valuelist: 需要插入的值数组

    """
    maxrow_index = sheet_.max_row
    for cell_index in range(1, len(valuelist) + 1):
        sheet_.cell(maxrow_index + 1, cell_index, valuelist[cell_index - 1])
    workbook_.save(filepath_)


def writerowfor2003(workbook_, filepath_, sheet_, valuelist):
    """
    在excel文件的最后插入一行数据.

    :param workbook_: 文件对象必须要formatting_info=True

    :param filepath_: 文件路径

    :param sheet_: excel的sheet对象

    :param valuelist: 需要插入的值数组

    """
    newwb = copy(workbook_)
    newsheet = newwb.get_sheet(0)
    for cell_index in range(0, len(valuelist)):
        if isinstance(valuelist[cell_index], datetime):
            newsheet.write(sheet_.nrows, cell_index, valuelist[cell_index].strftime('%Y/%m/%d %H:%M:%S'))
        else:
            newsheet.write(sheet_.nrows, cell_index, valuelist[cell_index])
    newwb.save(filepath_)


if __name__ == '__main__':  # 此为测试方法
    filepath_test = r'D:\11.xls'
    if len(filepath_test.split(r'.')) > 2:
        raise Exception('文件名非法%s' % filepath_test)
    elif filepath_test.split(r'.')[1] == 'xls':
        workbook_w = xlrd.open_workbook(filepath_test, formatting_info=True)
        sheet_w = workbook_w.sheet_by_index(0)
        writerowfor2003(workbook_w, filepath_test, sheet_w,
                        ['1', '2', '3', '2018/1/3  12:12:12', datetime.now()])
        workbook_r = xlrd.open_workbook(filepath_test, formatting_info=True)
        sheet_r = workbook_r.sheet_by_index(0)
        print(readexcelfor2003(workbook_r, sheet_r))
    elif filepath_test.split(r'.')[1] == 'xlsx':
        import openpyxl
        workbook_adv = openpyxl.load_workbook(filepath_test)
        sheet_adv = workbook_adv.active
        writerowfor2007(workbook_adv, filepath_test, sheet_adv, ['1', '2', '3', '2018/1/3  12:12:12', datetime.now()])
        print(readexcelfor2007(sheet_adv))
    else:
        raise Exception('请选择excel文件进行读取')
