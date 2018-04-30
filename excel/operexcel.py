"""本模块将所有操作excel的方法进行提供，并提供main方法调试"""

from datetime import datetime
import openpyxl
import xlrd
from xlutils.copy import copy


def read_excel(file_path_):
    """
    将excel封装成一个二位数组,，默认只读取第一个sheet.

    :param file_path_: excel的文件对象路径

    :return xls_list: 返回一个二维数组

    """
    xls_list = []  # 存储最终结果集
    temp_cell = {}  # 存储合并单元格的位置与值的字典
    if file_path_.split(r'.')[-1] == 'xls':
        workbook_2003 = xlrd.open_workbook(file_path_, formatting_info=True)
        sheet_2003 = workbook_2003.sheet_by_index(0)
        for merged_cell in sheet_2003.merged_cells:  # 将合并单元格的数据都赋值成同一个
            min_row, max_row, min_col, max_col = merged_cell
            for start_row_index in range(min_row, max_row):
                for endCol_index in range(min_col, max_col):
                    temp_cell[str(start_row_index) + str(endCol_index)] = sheet_2003.cell(min_row, min_col).value
        for row_index in range(sheet_2003.nrows):
            row_list = []  # 存储每一行的数据
            for col_index in range(sheet_2003.ncols):
                """ 这里判断当前单元格是否为空，为空是就取上个单元格的值，如果上一个也没值就取上上个，
                主要解决合并单元格的问题，但单元格本身为空时无法解决，下一次解决，xls是可以解决的。
                2018-4-29，肖志君修改将2003版本与2007版本分开读写都是               
                """
                if sheet_2003.cell(row_index, col_index).value == '':
                    if str(row_index)+str(col_index) in temp_cell.keys():
                        row_list.append(temp_cell.get(str(row_index)+str(col_index)))
                    else:
                        row_list.append('')
                else:
                    if sheet_2003.cell(row_index, col_index).ctype == 3:
                        date_value = xlrd.xldate_as_tuple(sheet_2003.cell(
                            row_index, col_index).value, workbook_2003.datemode)
                        date_tmp = datetime(*date_value).strftime('%Y/%m/%d %H:%M:%S')
                        row_list.append(date_tmp)
                    else:
                        row_list.append(sheet_2003.cell(row_index, col_index).value)
            xls_list.append(row_list)
    elif file_path_.split(r'.')[-1] == 'xlsx':
        workbook_2007 = openpyxl.load_workbook(file_path_test)
        sheet_2007 = workbook_2007.active
        for merged_cell in sheet_2007.merged_cells:
            for first_str in range(merged_cell.min_row, merged_cell.max_row + 1):
                for sedStr in range(merged_cell.min_col, merged_cell.max_col + 1):
                    temp_cell[str(first_str) + str(sedStr)] = \
                        sheet_2007.cell(merged_cell.min_row, merged_cell.min_col).value
        clo_num = 0  # 计算出第一行的列数，按此列数循环
        for row_ in sheet_2007.rows:
            clo_num = len(row_)
            break
        for row_index in range(sheet_2007.max_row):
            row_list = []  # 只取第一行的列数
            for clo_index in range(clo_num):
                if sheet_2007.cell(row_index + 1, clo_index + 1).value is None:  # 下标从一开始所以都要加1
                    if str(row_index) + str(row_index + 1) in temp_cell.keys():
                        row_list.append(temp_cell[str(row_index) + str(row_index + 1)])
                    else:
                        row_list.append('')
                else:
                    if isinstance(sheet_2007.cell(row_index + 1, clo_index + 1).value, datetime):
                        row_list.append(sheet_2007.cell(row_index + 1,
                                                        clo_index + 1).value.strftime('%Y/%m/%d %H:%M:%S'))
                    else:
                        row_list.append(sheet_2007.cell(row_index + 1, clo_index + 1).value)
            xls_list.append(row_list)
    else:
        raise FileNamedError()
    return xls_list


def write_excel(file_path_, value_list):
    """
    在excel文件的最后插入一行数据.

    :param file_path_: 文件路径

    :param value_list: 需要插入的值数组

    """
    if file_path_.split(r'.')[-1] == 'xlsx':
        workbook_2007 = openpyxl.load_workbook(file_path_test)
        sheet_2007 = workbook_2007.active
        max_row_index = sheet_2007.max_row
        for cell_index in range(1, len(value_list) + 1):
            sheet_2007.cell(max_row_index + 1, cell_index, value_list[cell_index - 1])
        workbook_2007.save(file_path_)
    elif file_path_.split(r'.')[-1] == 'xls':
        workbook_2003 = xlrd.open_workbook(file_path_, formatting_info=True)
        sheet_2003 = workbook_2003.sheet_by_index(0)
        new_wb = copy(workbook_2003)
        new_sheet = new_wb.get_sheet(0)
        for cell_index in range(0, len(value_list)):
            if isinstance(value_list[cell_index], datetime):
                new_sheet.write(sheet_2003.nrows, cell_index, value_list[cell_index].strftime('%Y/%m/%d %H:%M:%S'))
            else:
                new_sheet.write(sheet_2003.nrows, cell_index, value_list[cell_index])
        new_wb.save(file_path_)


class FileNamedError(Exception):
    def __init__(self, err='文件命名异常'):
        super(FileNamedError, self).__init__(self, err)


if __name__ == '__main__':  # 此为测试方法
    file_path_test = r'D:\11.xlsx'
    if len(file_path_test.split(r'.')) > 2:
        raise FileNamedError('文件名非法%s' % file_path_test)
    else:
        write_excel(file_path_test, ['1', '2', '3', '2018/1/3  12:12:12', datetime.now()])
        print(read_excel(file_path_test))
